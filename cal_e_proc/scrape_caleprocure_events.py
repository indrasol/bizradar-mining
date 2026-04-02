"""
Scrape Cal eProcure events from the California State Contracts Register.

Uses Playwright to:
  1. Log in to https://caleprocure.ca.gov
  2. Navigate to the Event Search page
  3. Collect event IDs from the search-results grid
  4. Open each event's detail page and extract structured fields
  5. Write everything to an Excel workbook
"""

import argparse
import re
import time
from dataclasses import dataclass

from playwright.sync_api import sync_playwright, Page
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# ── data model ────────────────────────────────────────────────────────────────

@dataclass
class EventRecord:
    event_id: str = ""
    title: str = ""
    dept: str = ""
    format_type: str = ""
    event_version: str = ""
    published_date: str = ""
    event_end_date: str = ""
    description: str = ""
    unspsc_classification: str = ""
    unspsc_classification_description: str = ""
    contractor_license_type: str = ""
    contractor_license_description: str = ""
    area_id: str = ""
    county: str = ""
    event_url: str = ""


EXCEL_HEADERS = [
    "Event ID",
    "Title",
    "Dept",
    "Format/Type",
    "Event Version",
    "Published Date",
    "Event End Date",
    "Description",
    "UNSPSC Classification",
    "UNSPSC Classification Description",
    "Contractor License Type",
    "Contractor License Description",
    "Area ID",
    "County",
    "Event URL",
]

SEARCH_URL = "https://caleprocure.ca.gov/pages/Events-BS3/event-search.aspx"
LOGIN_URL = "https://caleprocure.ca.gov/pages/BS3/login.aspx"


# ── helpers ───────────────────────────────────────────────────────────────────

def _clean(text: str) -> str:
    """Collapse whitespace, strip leading/trailing, replace &nbsp;."""
    text = text.replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def _wait_for_loading_done(page: Page, timeout: int = 30_000) -> None:
    """Wait for the InFlight loading overlay to disappear."""
    page.evaluate("""
        () => {
            const overlay = document.querySelector('#loadingContainer');
            if (overlay) overlay.style.display = 'none';
        }
    """)
    try:
        page.wait_for_selector(
            "#loadingContainer",
            state="hidden",
            timeout=timeout,
        )
    except Exception:
        page.evaluate("() => { const o = document.querySelector('#loadingContainer'); if (o) o.style.display = 'none'; }")


def _dismiss_modal(page: Page) -> None:
    """Close any modal dialog that might be blocking interaction."""
    page.evaluate("""
        () => {
            const modal = document.querySelector('#popupMessageModal');
            if (modal) {
                modal.style.display = 'none';
                modal.classList.remove('in');
                modal.setAttribute('aria-hidden', 'true');
            }
            const backdrop = document.querySelector('.modal-backdrop');
            if (backdrop) backdrop.remove();
            document.body.classList.remove('modal-open');
            document.body.style.overflow = '';
            document.body.style.paddingRight = '';
        }
    """)


def _is_logged_in(page: Page) -> bool:
    """Return True if the current page is still in an authenticated session."""
    url = page.url.lower()
    if "login.aspx" in url:
        return False
    return page.evaluate("""
        () => {
            const logoutEl = document.querySelector('#logout, #logout_tpl[style*="display"]');
            const userName = document.querySelector('#user-name');
            if (userName && userName.textContent.trim()) return true;
            if (logoutEl) return true;
            const loginLink = document.querySelector('#sign-in-link');
            if (loginLink) {
                const style = window.getComputedStyle(loginLink.closest('#login_tpl') || loginLink);
                if (style.display !== 'none') return false;
            }
            return true;
        }
    """)


def login(page: Page, user_id: str, password: str) -> None:
    """Log in to Cal eProcure."""
    page.goto(LOGIN_URL, wait_until="domcontentloaded")
    page.wait_for_selector("#userid", timeout=30_000)
    page.fill("#userid", user_id)
    page.fill("#pwd", password)
    page.click("#submitButton")
    page.wait_for_url("**/index.aspx**", timeout=30_000)
    print("  Logged in successfully.")


def _ensure_search_page(page: Page, user_id: str, password: str) -> None:
    """Make sure we are logged in, on the search page, AND the form is usable."""
    _close_extra_tabs(page)

    logged_in_1 = _is_logged_in(page)

    if not logged_in_1:
        print("    (session expired, re-logging in...)")
        login(page, user_id, password)

    current = page.url.lower()
    if "event-search" not in current:
        page.goto(SEARCH_URL, wait_until="domcontentloaded")

    if not _is_logged_in(page):
        print("    (session expired on navigate, re-logging in...)")
        login(page, user_id, password)
        page.goto(SEARCH_URL, wait_until="domcontentloaded")

    page.wait_for_timeout(2_000)
    _wait_for_loading_done(page)
    _dismiss_modal(page)

    # Verify the form is actually rendered (InFlight may have died silently)
    if not _is_search_form_ready(page):
        print("    (search form not rendered, reloading page...)")
        page.reload(wait_until="domcontentloaded")
        page.wait_for_timeout(5_000)
        _wait_for_loading_done(page)
        _dismiss_modal(page)

        if not _is_search_form_ready(page):
            print("    (form still missing after reload, full re-login...)")
            login(page, user_id, password)
            page.goto(SEARCH_URL, wait_until="domcontentloaded")
            page.wait_for_timeout(5_000)
            _wait_for_loading_done(page)
            _dismiss_modal(page)


def _is_search_form_ready(page: Page, wait_ms: int = 8_000) -> bool:
    """Check if the Event ID textbox is actually present in the DOM."""
    try:
        page.get_by_role("textbox", name="Event ID").wait_for(state="visible", timeout=wait_ms)
        return True
    except Exception:
        return False


def _close_extra_tabs(page: Page) -> None:
    """Close all tabs except the one we're working on."""
    context = page.context
    for p in context.pages:
        if p != page:
            try:
                p.close()
            except Exception:
                pass


def collect_event_ids(page: Page, max_events: int) -> list[str]:
    """
    Navigate to the search page and collect up to *max_events* Event IDs
    from the default "Posted" search results grid.
    """
    page.goto(SEARCH_URL, wait_until="domcontentloaded")

    # Wait for the "Searching..." overlay to disappear and the grid to render
    page.wait_for_selector(
        "#datatable-ready tbody tr",
        state="attached",
        timeout=60_000,
    )
    # Small extra wait for the InFlight JS to finish populating the grid
    page.wait_for_timeout(2_000)

    event_ids: list[str] = page.evaluate("""
        () => {
            const rows = document.querySelectorAll('#datatable-ready tbody tr');
            const ids = [];
            for (const row of rows) {
                const cell = row.querySelector('td[data-if-label="tdEventId"]');
                if (cell) ids.push(cell.textContent.trim());
            }
            return ids;
        }
    """)

    result = event_ids[:max_events] if max_events > 0 else event_ids
    print(f"  Collected {len(result)} event IDs from search grid.")
    return result


def open_event_and_extract(
    search_page: Page, event_id: str, user_id: str, password: str,
) -> EventRecord:
    """
    From the search page, search for *event_id*, click the row to open its
    detail tab, then scrape all required fields.  Retries up to twice on failure.
    """
    max_attempts = 5
    for attempt in range(max_attempts):
        try:
            rec = _try_open_and_extract(search_page, event_id, user_id, password)
            return rec
        except Exception as exc:
            if attempt < max_attempts - 1:
                print(f"    RETRY ({attempt+1}): {type(exc).__name__}: {str(exc)[:80]}", flush=True)
                search_page.bring_to_front()
                _ensure_search_page(search_page, user_id, password)
            else:
                print(f"    WARN: Could not extract {event_id}: {str(exc)[:100]}", flush=True)
                return EventRecord(event_id=event_id)

    return EventRecord(event_id=event_id)


def _try_open_and_extract(
    search_page: Page, event_id: str, user_id: str, password: str,
) -> EventRecord:
    """Single attempt to search, open, and extract an event."""
    search_page.bring_to_front()

    # Always verify session before proceeding
    _ensure_search_page(search_page, user_id, password)

    _fill_event_id_and_search(search_page, event_id)

    detail_page = _click_event_row(search_page, event_id)
    if detail_page is None:
        raise RuntimeError(f"Could not open detail page for {event_id}")

    try:
        record = _extract_detail(detail_page, event_id)
    finally:
        detail_page.close()

    return record


def _fill_event_id_and_search(page: Page, event_id: str) -> None:
    """Type an event ID into the search box and click Search."""
    _wait_for_loading_done(page)
    _dismiss_modal(page)

    eid_box = page.get_by_role("textbox", name="Event ID")
    eid_box.fill(event_id)

    _wait_for_loading_done(page)
    _dismiss_modal(page)

    page.locator("#RESP_INQA_WK_INQ_AUC_GO_PB").click(force=True)

    # Wait for the grid to refresh
    page.wait_for_timeout(2_000)
    _wait_for_loading_done(page)
    _dismiss_modal(page)
    page.wait_for_selector("#datatable-ready tbody tr", state="attached", timeout=30_000)
    page.wait_for_timeout(1_000)


def _click_event_row(search_page: Page, event_id: str) -> Page | None:
    """Click the grid row for *event_id* and return the newly-opened Page."""
    context = search_page.context

    _wait_for_loading_done(search_page)
    _dismiss_modal(search_page)

    with context.expect_page(timeout=30_000) as new_page_info:
        search_page.evaluate("""
            (eid) => {
                const cells = document.querySelectorAll(
                    'td[data-if-label="tdEventId"]'
                );
                for (const td of cells) {
                    if (td.textContent.trim() === eid) {
                        td.click();
                        return;
                    }
                }
            }
        """, event_id)

    detail_page = new_page_info.value
    _wait_for_detail_ready(detail_page)
    return detail_page


def _wait_for_detail_ready(page: Page, timeout: int = 60_000) -> None:
    """Wait for the detail page to finish loading and InFlight to populate.
    Raises RuntimeError if the page stays on loading.html or content never appears."""
    # Phase 1: wait for URL to leave loading.html
    url_ok = False
    try:
        page.wait_for_url("**/event/**", timeout=timeout)
        url_ok = True
    except Exception:
        try:
            page.wait_for_url("**/event-details**", timeout=15_000)
            url_ok = True
        except Exception:
            pass

    current_url = page.url
    if not url_ok or "loading.html" in current_url:
        raise RuntimeError(f"Detail page stuck on loading.html: {current_url}")

    page.wait_for_load_state("domcontentloaded")
    page.wait_for_timeout(3_000)

    # Phase 2: wait for InFlight to inject real content
    deadline = time.time() + 45
    while time.time() < deadline:
        ready = page.evaluate("""
            () => {
                const nameEl = document.querySelector('[data-if-label="eventName"]');
                const idEl   = document.querySelector('[data-if-label="eventId"]');
                const name = nameEl ? nameEl.textContent.trim() : '';
                const id   = idEl   ? idEl.textContent.trim()   : '';
                if (name && name !== '[Event Title]' && id) return true;
                return false;
            }
        """)
        if ready:
            return
        page.wait_for_timeout(1_500)

    raise RuntimeError(f"InFlight content never populated on {page.url}")


def _extract_detail(page: Page, event_id: str) -> EventRecord:
    """Pull every field from the event-detail page."""

    data: dict = page.evaluate("""
        () => {
            const txt = (sel) => {
                const el = document.querySelector(sel);
                return el ? el.textContent.trim() : '';
            };
            const attr = (sel, a) => {
                const el = document.querySelector(sel);
                return el ? (el.getAttribute(a) || '') : '';
            };

            // Title
            const title = txt('[data-if-label="eventName"]')
                       || txt('h3.bold[data-if-label="eventName"]')
                       || txt('h3.bold');

            // Event ID
            const eventId = txt('[data-if-label="eventId"]')
                         || txt('[data-if-source="#RESP_AUC_H0B_WK_AUC_ID_BUS_UNIT"]');

            // Format / Type
            const fmt1 = txt('[data-if-label="format1"]');
            const fmt2 = txt('[data-if-label="format2"]');
            const formatType = [fmt1, fmt2].filter(Boolean).join(' / ');

            // Published date
            const pubDate = txt('[data-if-label="eventStartDate"]');

            // Event End Date
            const endDate = txt('[data-if-label="eventEndDate"]');

            // Dept
            const dept = txt('[data-if-label="dept"]');

            // Event Version
            const version = txt('[data-if-label="eventVersion"]');

            // Description
            const desc = txt('[data-if-label="descriptiondetails"]');

            // UNSPSC table
            const unspscRows = document.querySelectorAll(
                '#unspscCodesSection ~ thead ~ tbody tr, '
                + 'table:has(#unspscCodesSection) tbody tr'
            );
            const unspscCodes = [];
            const unspscDescs = [];
            for (const row of unspscRows) {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 2) {
                    const code = cells[0].textContent.trim();
                    const d    = cells[1].textContent.trim();
                    if (code && code !== '\u00a0') { unspscCodes.push(code); unspscDescs.push(d); }
                }
            }

            // Contractor License Type table
            const clRows = document.querySelectorAll(
                '#contractorLicenseTypeSection ~ thead ~ tbody tr, '
                + 'table:has(#contractorLicenseTypeSection) tbody tr, '
                + '#contractorTable tbody tr'
            );
            const clTypes = [];
            const clDescs = [];
            for (const row of clRows) {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 2) {
                    const t = cells[0].textContent.trim();
                    const d = cells[1].textContent.trim();
                    if (t && t !== '\u00a0') { clTypes.push(t); clDescs.push(d); }
                }
            }

            // Service Area table
            const saRows = document.querySelectorAll(
                '#serviceAreaSection ~ thead ~ tbody tr, '
                + 'table:has(#serviceAreaSection) tbody tr, '
                + '#serviceAreaTable tbody tr'
            );
            const areaIds  = [];
            const counties = [];
            for (const row of saRows) {
                const cells = row.querySelectorAll('td');
                if (cells.length >= 2) {
                    const a = cells[0].textContent.trim();
                    const c = cells[1].textContent.trim();
                    if (a && a !== '\u00a0') { areaIds.push(a); counties.push(c); }
                }
            }

            // Event URL from the share text box
            const urlInput = document.querySelector('#shareEventText');
            const eventUrl = urlInput ? urlInput.value : window.location.href;

            return {
                eventId,
                title,
                dept,
                formatType,
                version,
                pubDate,
                endDate,
                desc,
                unspscCodes:  unspscCodes.join('; '),
                unspscDescs:  unspscDescs.join('; '),
                clTypes:      clTypes.join('; '),
                clDescs:      clDescs.join('; '),
                areaIds:      areaIds.join('; '),
                counties:     counties.join('; '),
                eventUrl,
            };
        }
    """)

    return EventRecord(
        event_id=_clean(data.get("eventId", event_id)),
        title=_clean(data.get("title", "")),
        dept=_clean(data.get("dept", "")),
        format_type=_clean(data.get("formatType", "")),
        event_version=_clean(data.get("version", "")),
        published_date=_clean(data.get("pubDate", "")),
        event_end_date=_clean(data.get("endDate", "")),
        description=_clean(data.get("desc", "")),
        unspsc_classification=data.get("unspscCodes", ""),
        unspsc_classification_description=data.get("unspscDescs", ""),
        contractor_license_type=data.get("clTypes", "") or "N/A",
        contractor_license_description=data.get("clDescs", "") or "N/A",
        area_id=data.get("areaIds", "") or "N/A",
        county=data.get("counties", "") or "N/A",
        event_url=data.get("eventUrl", ""),
    )


# ── Excel output ──────────────────────────────────────────────────────────────

def write_excel(records: list[EventRecord], path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cal eProcure Events"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, rec in enumerate(records, start=2):
        values = [
            rec.event_id,
            rec.title,
            rec.dept,
            rec.format_type,
            rec.event_version,
            rec.published_date,
            rec.event_end_date,
            rec.description,
            rec.unspsc_classification,
            rec.unspsc_classification_description,
            rec.contractor_license_type,
            rec.contractor_license_description,
            rec.area_id,
            rec.county,
            rec.event_url,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap

    col_widths = [16, 45, 30, 22, 14, 22, 22, 60, 22, 50, 24, 40, 12, 20, 55]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"\n  Excel saved -> {path}")


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Scrape Cal eProcure events to Excel")
    parser.add_argument("--user", default="Indrasol", help="Cal eProcure User ID")
    parser.add_argument("--password", default="btcA8tT@Gh33x1l@", help="Cal eProcure password")
    parser.add_argument("--max-events", type=int, default=5, help="Number of events to scrape")
    parser.add_argument("--output", default="caleprocure_events.xlsx", help="Output Excel file path")
    parser.add_argument("--headless", action="store_true", help="Run browser in headless mode")
    parser.add_argument("--resume-from", type=int, default=0, help="Skip the first N events (resume from checkpoint)")
    args = parser.parse_args()

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=args.headless)
        context = browser.new_context(viewport={"width": 1440, "height": 900})
        page = context.new_page()

        # Step 1 – Login
        print("[1/3] Logging in ...")
        login(page, args.user, args.password)

        # Step 2 – Collect event IDs from search results
        print("[2/3] Collecting event IDs from search grid ...")
        event_ids = collect_event_ids(page, args.max_events)

        # Step 3 – Open each event and extract details
        if args.resume_from > 0:
            event_ids = event_ids[args.resume_from:]
            print(f"  Resuming from offset {args.resume_from}, {len(event_ids)} events remaining.")

        total = len(event_ids)
        print(f"[3/3] Extracting details for {total} events ...")
        records: list[EventRecord] = []
        failed_ids: list[str] = []
        run_start = time.time()
        for i, eid in enumerate(event_ids, start=1):
            print(f"  ({i}/{total}) Event {eid} ...", flush=True)
            rec = open_event_and_extract(page, eid, args.user, args.password)
            records.append(rec)
            if not rec.title:
                failed_ids.append(eid)
            print(f"    OK: {rec.title[:60]}", flush=True)

            # Incremental save every 25 events (H4)
            if i % 25 == 0:
                write_excel(records, args.output)
                elapsed = time.time() - run_start
                rate = i / elapsed * 60 if elapsed > 0 else 0
                print(f"  [checkpoint] {i}/{total} saved, {len(failed_ids)} failed, {rate:.0f} events/min", flush=True)

        browser.close()

    # Final save
    write_excel(records, args.output)
    elapsed = time.time() - run_start
    print(f"\nDone - scraped {len(records)} events ({len(records)-len(failed_ids)} OK, {len(failed_ids)} failed) in {elapsed/60:.1f} min.")
    if failed_ids:
        print(f"Failed event IDs: {', '.join(failed_ids[:30])}{'...' if len(failed_ids) > 30 else ''}")


if __name__ == "__main__":
    main()
